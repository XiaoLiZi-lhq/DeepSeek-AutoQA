"""
P4: Excel 读取 + 数据校验
提供 ExcelReader 类，读取包含 brand_name 和 question 列的问题集。
"""

import os
from typing import Any

import openpyxl


class ExcelReader:
    """读取并校验包含品牌问题集的 Excel 文件。"""

    REQUIRED_COLUMNS = {"brand_name", "question"}

    def __init__(self, file_path: str):
        """
        初始化 ExcelReader。

        Args:
            file_path: Excel 文件的绝对或相对路径。
        """
        self.file_path = file_path

    # ------------------------------------------------------------------
    # 校验方法
    # ------------------------------------------------------------------

    def validate(self) -> bool:
        """
        验证 Excel 文件是否可用。

        检查项：
        - 文件存在
        - 可正常打开（不是损坏文件）
        - 包含 brand_name 和 question 列
        - 至少有一条有效数据

        Returns:
            True 表示校验通过，False 表示存在致命问题。
        """
        if not os.path.isfile(self.file_path):
            print(f"[警告] 文件不存在: {self.file_path}")
            return False

        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
        except Exception as e:
            print(f"[警告] 无法打开 Excel 文件（可能已损坏）: {e}")
            return False

        try:
            ws = wb.active
            if ws is None:
                print("[警告] Excel 中没有工作表")
                return False

            header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            missing = self.REQUIRED_COLUMNS - set(str(h or "").strip() for h in header_row)
            if missing:
                print(f"[警告] Excel 缺少必要列: {missing}，表头为 {header_row}")
                return False

            col_indices = {}
            for idx, val in enumerate(header_row):
                if val in self.REQUIRED_COLUMNS:
                    col_indices[str(val)] = idx

            has_data = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                brand = row[col_indices["brand_name"]] if len(row) > col_indices["brand_name"] else None
                question = row[col_indices["question"]] if len(row) > col_indices["question"] else None
                if brand and str(brand).strip() and question and str(question).strip():
                    has_data = True
                    break

            if not has_data:
                print("[警告] Excel 中没有有效数据行")
                return False

            return True

        finally:
            wb.close()

    # ------------------------------------------------------------------
    # 读取方法
    # ------------------------------------------------------------------

    def read_questions(self) -> list[dict[str, str]]:
        """
        读取 Excel 中的问题列表。

        Excel 格式要求：首行为表头，必须包含 brand_name 和 question 两列。
        异常行处理：空值行、brand_name 或 question 为空的行自动跳过并打印警告。

        Returns:
            问题列表，每项为 {"brand_name": "华为", "question": "华为2024年营收？"}
        """
        if not os.path.isfile(self.file_path):
            print(f"[错误] 文件不存在: {self.file_path}")
            return []

        wb = openpyxl.load_workbook(self.file_path, read_only=True)
        questions: list[dict[str, str]] = []
        total_rows = 0
        skipped_rows = 0

        try:
            ws = wb.active
            if ws is None:
                print("[错误] Excel 中没有工作表")
                return []

            header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            col_map: dict[str, int] = {}
            for idx, val in enumerate(header_row):
                if val in self.REQUIRED_COLUMNS:
                    col_map[str(val)] = idx

            missing = self.REQUIRED_COLUMNS - set(col_map.keys())
            if missing:
                print(f"[错误] Excel 缺少必要列: {missing}，表头为 {header_row}")
                return []

            brand_idx = col_map["brand_name"]
            question_idx = col_map["question"]

            for row in ws.iter_rows(min_row=2, values_only=True):
                total_rows += 1
                brand = row[brand_idx] if len(row) > brand_idx else None
                question = row[question_idx] if len(row) > question_idx else None

                brand_str = str(brand).strip() if brand is not None else ""
                question_str = str(question).strip() if question is not None else ""

                if not brand_str and not question_str:
                    print(f"[警告] 第 {total_rows + 1} 行（数据行 {total_rows}）完全为空，已跳过")
                    skipped_rows += 1
                    continue

                if not brand_str:
                    print(f"[警告] 第 {total_rows + 1} 行（数据行 {total_rows}）brand_name 为空，已跳过")
                    skipped_rows += 1
                    continue

                if not question_str:
                    print(f"[警告] 第 {total_rows + 1} 行（数据行 {total_rows}）question 为空，已跳过")
                    skipped_rows += 1
                    continue

                questions.append({
                    "brand_name": brand_str,
                    "question": question_str,
                })

        finally:
            wb.close()

        valid_rows = len(questions)
        print(f"Excel 读取完毕: 总行数={total_rows}, 有效行数={valid_rows}, 跳过行数={skipped_rows}")
        return questions
